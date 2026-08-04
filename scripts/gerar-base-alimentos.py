#!/usr/bin/env python
"""Gera a base nutricional v1 do Equivale a partir da planilha oficial TACO.

Dependência: openpyxl==3.1.5. O script não acessa a rede e nunca consulta a TBCA.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
import unicodedata
from collections import defaultdict
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
TACO = ROOT / "data/fontes/taco/Taco-4a-Edicao.xlsx"
CATALOGO = ROOT / "incoming/data/catalogo-equivalencias-equivale.md"
BANCO_GERAL = ROOT / "incoming/data/equivale-food-database.md"
MODELO = ROOT / "src/importacao/modelo.ts"
SELECAO = ROOT / "data/alimentos-v1-selecao.json"
SAIDA = ROOT / "public/data/alimentos.json"
RELATORIO_JSON = ROOT / "docs/alimentos-v1-relatorio.json"
RELATORIO_MD = ROOT / "docs/alimentos-v1-relatorio.md"
INVENTARIO_JSON = ROOT / "docs/alimentos-inventario-pre-base.json"
INVENTARIO_MD = ROOT / "docs/alimentos-inventario-pre-base.md"
ACESSO_EM = "2026-08-03"
FONTE = "TACO — Tabela Brasileira de Composição de Alimentos, 4ª edição ampliada e revisada (NEPA/Unicamp)"
URL = "https://www.nepa.unicamp.br/wp-content/uploads/sites/27/2023/10/Taco-4a-Edicao.xlsx"
SHA256 = "a66b8ec528daeabc63bc2b015fc9bd8c6d76b941c2fc0ed93a4311d449302d14"

CATEGORIAS = {
    "Cereais e derivados": "cereais-paes-massas-tuberculos",
    "Verduras, hortaliças e derivados": "hortalicas",
    "Frutas e derivados": "frutas",
    "Gorduras e óleos": "gorduras",
    "Pescados e frutos do mar": "proteinas",
    "Carnes e derivados": "proteinas",
    "Leite e derivados": "laticinios",
    "Bebidas (alcoólicas e não alcoólicas)": "bebidas",
    "Ovos e derivados": "proteinas",
    "Produtos açucarados": "acucares-e-doces",
    "Miscelâneas": "industrializados",
    "Outros alimentos industrializados": "industrializados",
    "Alimentos preparados": "receitas",
    "Leguminosas e derivados": "leguminosas",
    "Nozes e sementes": "gorduras",
}
CATEGORIAS_BANCO_GERAL = {
    "aves": "proteinas", "bebidas": "bebidas", "carnes bovinas": "proteinas",
    "cereais": "cereais-paes-massas-tuberculos", "comida japonesa": "receitas",
    "condimentos": "industrializados", "doces": "acucares-e-doces", "frutas": "frutas",
    "hambúrgueres": "proteinas", "hortaliças": "hortalicas",
    "industrializados": "industrializados", "ingredientes culinários": "industrializados",
    "laticínios": "laticinios", "leguminosas": "leguminosas",
    "massas": "cereais-paes-massas-tuberculos", "outras carnes": "proteinas",
    "ovos": "proteinas", "peixes e frutos do mar": "proteinas", "pizzas": "receitas",
    "pratos preparados": "receitas", "produtos de marca": "industrializados",
    "pães": "cereais-paes-massas-tuberculos", "salgados": "receitas",
    "sanduíches": "receitas", "sobremesas": "acucares-e-doces", "suínos": "proteinas",
    "tubérculos": "cereais-paes-massas-tuberculos", "óleos": "gorduras",
}
ALIASES_CONTROLADOS = {
    "taco-004-arroz-tipo-1-cru": ["Arroz branco cru"],
    "taco-038-lasanha-massa-fresca-crua": ["Massa de lasanha crua", "Massa de lasanha fresca crua"],
    "taco-040-macarrao-trigo-cru": ["Macarrão seco", "Massa seca de macarrão"],
}
PREPAROS = (
    ("cozida/10minutos", "cozido"), ("cozido/10minutos", "cozido"),
    ("grelhada", "grelhado"), ("grelhado", "grelhado"),
    ("assada", "assado"), ("assado", "assado"),
    ("refogada", "refogado"), ("refogado", "refogado"),
    ("frita", "frito"), ("frito", "frito"),
    ("cozida", "cozido"), ("cozido", "cozido"),
    ("crua", "cru"), ("cru", "cru"),
    ("drenada", "drenado"), ("drenado", "drenado"),
)


def normalizar(texto: str) -> str:
    texto = "".join(c for c in unicodedata.normalize("NFD", texto.lower()) if unicodedata.category(c) != "Mn")
    return re.sub(r"[^a-z0-9]+", " ", texto).strip()


def slug(texto: str) -> str:
    return normalizar(texto).replace(" ", "-")


def ler_catalogo() -> list[dict]:
    registros = []
    for linha in CATALOGO.read_text(encoding="utf-8").splitlines():
        achado = re.match(r"^\s*-\s*(\{.*\})\s*$", linha)
        if not achado:
            continue
        try:
            item = json.loads(achado.group(1))
        except json.JSONDecodeError:
            continue
        if item.get("nomePadronizado"):
            registros.append(item)
    return registros


def ler_modelo() -> list[dict]:
    registros = []
    for linha in MODELO.read_text(encoding="utf-8").splitlines():
        if linha.startswith("- ") and "|" in linha:
            nome = linha[2:].split("|", 1)[0].strip()
            registros.append({"nomePadronizado": nome, "nomeOriginal": nome, "origemDado": "modelo-markdown", "planosEmQueAparece": ["modelo"]})
    return registros


def ler_taco() -> dict[str, dict]:
    if hashlib.sha256(TACO.read_bytes()).hexdigest() != SHA256:
        raise ValueError("O SHA-256 da planilha TACO não corresponde ao arquivo oficial auditado.")
    planilha = load_workbook(TACO, data_only=True, read_only=True).active
    categoria = None
    itens = {}
    for linha in planilha.iter_rows(values_only=True):
        primeiro, descricao = linha[0], linha[1]
        if isinstance(primeiro, str) and primeiro.strip() in CATEGORIAS:
            categoria = CATEGORIAS[primeiro.strip()]
            continue
        if not isinstance(primeiro, (int, float)) or not isinstance(descricao, str):
            continue
        kcal = linha[3]
        codigo = str(int(primeiro))
        itens[codigo] = {"codigo": codigo, "descricao": descricao.strip(), "kcal": kcal, "categoria": categoria}
    return itens


def preparo_da_descricao(descricao: str) -> str | None:
    chave = normalizar(descricao)
    for sufixo, preparo in PREPAROS:
        if chave.endswith(normalizar(sufixo)):
            return preparo
    return None


def limpar_medida(texto: str) -> str:
    texto = re.sub(r"\((?:e)?s\)", "", texto, flags=re.I)
    texto = texto.replace("bífe", "bife").replace("filé", "file")
    texto = re.sub(r"\s+", " ", texto).strip().lower()
    # O parser da aplicação normaliza acentos; mantemos forma legível e singular documentada.
    plurais = {"colheres": "colher", "unidades": "unidade", "fatias": "fatia", "pedaços": "pedaço", "asas": "asa"}
    partes = texto.split(" ")
    if partes and partes[0] in plurais:
        partes[0] = plurais[partes[0]]
    return " ".join(partes)


def porcoes_documentadas(registros: list[dict]) -> list[dict]:
    porcoes = []
    vistas = set()
    for item in registros:
        medida = item.get("medidaCaseira")
        quantidade = item.get("quantidadeMedidaCaseira")
        gramas = item.get("pesoGramas")
        if not medida or not isinstance(quantidade, (int, float)) or not isinstance(gramas, (int, float)):
            continue
        if quantidade <= 0 or gramas <= 0 or "REVIEW_REQUIRED" in str(item.get("observacoes") or ""):
            continue
        medida_limpa = limpar_medida(medida)
        chave = (normalizar(medida_limpa), float(quantidade), float(gramas))
        if chave in vistas:
            continue
        vistas.add(chave)
        porcoes.append({
            "medida": medida_limpa,
            "quantidade": float(quantidade),
            "gramas": float(gramas),
            "fonte": item.get("referencia", "Catálogo de substituições da nutricionista"),
            "codigoDaFonte": item.get("id"),
        })
    return porcoes


def ler_banco_geral() -> list[dict]:
    """Lê os registros operacionais JSON do Markdown recebido."""
    registros = []
    for linha in BANCO_GERAL.read_text(encoding="utf-8").splitlines():
        achado = re.match(r"^\s*-\s*(\{.*\})\s*$", linha)
        if not achado:
            continue
        try:
            item = json.loads(achado.group(1))
        except json.JSONDecodeError:
            continue
        if item.get("tipoRegistro") in {
            "alimento-individual", "ingrediente", "preparacao-generica", "produto-de-marca"
        }:
            registros.append(item)
    return registros


def converter_banco_geral(enriquecimentos_taco: list[dict], taco: dict[str, dict]) -> list[dict]:
    """Converte o banco geral para o contrato enxuto consumido pelo aplicativo."""
    por_codigo_taco = {item["codigoDaFonte"]: item for item in enriquecimentos_taco}
    alimentos = []
    for item in ler_banco_geral():
        kcal = item.get("kcalPor100g")
        if item.get("utilizavelEmEquivalenciaCalorica") is not True and item.get("tipoRegistro") != "produto-de-marca":
            continue
        if not isinstance(kcal, (int, float)) or not math.isfinite(kcal) or kcal < 0:
            continue

        codigo_bruto = str(item.get("codigoFonte") or "")
        codigo = codigo_bruto.removeprefix("TACO:")
        if codigo.isdigit():
            codigo = str(int(codigo))
        anterior = por_codigo_taco.get(codigo)
        origem_taco = taco.get(codigo)
        aliases = [*(item.get("aliases") or []), *ALIASES_CONTROLADOS.get(item["id"], [])]
        porcoes = []
        for porcao in item.get("medidasCaseirasDocumentadas") or []:
            if not isinstance(porcao, dict):
                continue
            quantidade, gramas = porcao.get("quantidade"), porcao.get("gramas")
            medida = porcao.get("unidade") or porcao.get("descricao")
            if not isinstance(medida, str) or not isinstance(quantidade, (int, float)) or not isinstance(gramas, (int, float)):
                continue
            if quantidade <= 0 or gramas <= 0:
                continue
            porcoes.append({
                "medida": limpar_medida(medida), "quantidade": float(quantidade),
                "gramas": float(gramas),
                "fonte": porcao.get("referencia") or porcao.get("fonte") or item.get("fonte"),
                "codigoDaFonte": porcao.get("codigoFonte"),
            })
        if anterior:
            aliases.extend(anterior["aliases"])
            porcoes.extend(anterior["porcoesCaseiras"])

        peso_porcao = item.get("pesoPorcaoGramas")
        if item.get("tipoRegistro") == "produto-de-marca" and isinstance(peso_porcao, (int, float)) and peso_porcao > 0:
            porcoes.append({
                "medida": "unidade", "quantidade": 1.0, "gramas": float(peso_porcao),
                "fonte": item.get("fonte"), "codigoDaFonte": codigo_bruto,
            })

        preparos = item.get("preparo") or []
        preparo = ", ".join(preparos) if isinstance(preparos, list) and preparos else None
        porcoes_unicas = {(p["medida"], p["quantidade"], p["gramas"]): p for p in porcoes}
        alimentos.append({
            "id": item["id"], "nome": item.get("nomeCanonico") or item.get("nomeProduto"),
            "aliases": sorted(set(aliases)),
            "categoria": origem_taco["categoria"] if origem_taco else CATEGORIAS_BANCO_GERAL.get(item.get("categoria"), "industrializados"),
            "preparo": preparo, "kcalPor100g": round(float(kcal), 6),
            "porcoesCaseiras": list(porcoes_unicas.values()),
            "fonte": item.get("fonte") or "não informada", "codigoDaFonte": codigo or item["id"],
            "atualizadoEm": item.get("dataAtualizacao") or item.get("dataAcesso") or "",
        })
    return alimentos


def gerar() -> tuple[dict, dict]:
    taco = ler_taco()
    catalogo = ler_catalogo()
    inventario = catalogo + ler_modelo()
    por_nome = defaultdict(list)
    for item in inventario:
        por_nome[normalizar(item["nomePadronizado"])].append(item)

    taco_por_nome = {normalizar(item["descricao"]): codigo for codigo, item in taco.items()}
    correspondencias = {}
    for nome in por_nome:
        if nome in taco_por_nome:
            correspondencias[nome] = taco_por_nome[nome]
    selecao = json.loads(SELECAO.read_text(encoding="utf-8"))
    for mapa in selecao["mapeamentos"]:
        chave = normalizar(mapa["nome"])
        if chave not in por_nome:
            raise ValueError(f"Mapeamento não existe no inventário: {mapa['nome']}")
        correspondencias[chave] = mapa["codigoTaco"]

    registros_por_codigo = defaultdict(list)
    for chave, codigo in correspondencias.items():
        registros_por_codigo[codigo].extend(por_nome[chave])

    alimentos = []
    invalidos = []
    for codigo in sorted(registros_por_codigo, key=lambda valor: int(valor)):
        origem = taco.get(codigo)
        if not origem or not isinstance(origem["kcal"], (int, float)) or not math.isfinite(origem["kcal"]) or origem["kcal"] < 0:
            invalidos.append({"codigo": codigo, "motivo": "kcal ausente ou inválida"})
            continue
        associados = registros_por_codigo[codigo]
        aliases = sorted({item.get("nomePadronizado") for item in associados if item.get("nomePadronizado") and normalizar(item["nomePadronizado"]) != normalizar(origem["descricao"])})
        alimentos.append({
            "id": f"taco-{codigo}",
            "nome": origem["descricao"],
            "aliases": aliases,
            "categoria": origem["categoria"],
            "preparo": preparo_da_descricao(origem["descricao"]),
            "kcalPor100g": round(float(origem["kcal"]), 6),
            "porcoesCaseiras": porcoes_documentadas([item for item in associados if item.get("origemDado") != "modelo-markdown"]),
            "fonte": FONTE,
            "codigoDaFonte": codigo,
            "atualizadoEm": ACESSO_EM,
        })

    alimentos = converter_banco_geral(alimentos, taco)
    base = {
        "versao": 2,
        "fonte": "Banco geral do Equivale — TACO 4ª edição e fontes oficiais identificadas",
        "somenteParaTeste": False,
        "alimentos": alimentos,
    }
    chaves = defaultdict(set)
    for alimento in alimentos:
        for texto in [alimento["nome"], *alimento["aliases"]]:
            chaves[normalizar(texto)].add(alimento["id"])
    conflitos = [{"chave": chave, "ids": sorted(ids)} for chave, ids in chaves.items() if len(ids) > 1]
    nomes_resolvidos = set(correspondencias)
    sem_correspondencia = []
    resolviveis = []
    for chave, itens in por_nome.items():
        if chave not in nomes_resolvidos:
            sem_correspondencia.append(sorted({i["nomePadronizado"] for i in itens})[0])
            continue
        codigo = correspondencias[chave]
        if any(isinstance(i.get("pesoGramas"), (int, float)) and i["pesoGramas"] > 0 for i in itens) and codigo not in {x["codigo"] for x in invalidos}:
            resolviveis.append(sorted({i["nomePadronizado"] for i in itens})[0])
    relatorio = {
        "geradoEm": ACESSO_EM,
        "fontePrimaria": {"publicacao": FONTE, "url": URL, "sha256": SHA256, "acessadoEm": ACESSO_EM},
        "fonteComplementar": None,
        "tbcaNaoUtilizada": "A versão 7.3 informa CC BY-NC-ND 4.0 e proíbe reprodução total ou parcial; nenhuma página individual foi coletada.",
        "alimentosGerados": len(alimentos),
        "registrosTacoDisponiveis": len(taco),
        "ocorrenciasInventariadas": len(inventario),
        "nomesNormalizadosInventariados": len(por_nome),
        "duplicatasNormalizadas": sum(1 for itens in por_nome.values() if len(itens) > 1),
        "aliasesConflitantes": conflitos,
        "valoresInvalidosOuAusentesExcluidos": invalidos,
        "medidasCaseirasDisponiveis": sum(len(a["porcoesCaseiras"]) for a in alimentos),
        "itensComPesoDocumentadoResolvíveis": sorted(set(resolviveis)),
        "quantidadeItensComPesoDocumentadoResolvíveis": len(set(resolviveis)),
        "alimentosAindaSemCorrespondencia": sorted(set(sem_correspondencia)),
        "quantidadeAindaSemCorrespondencia": len(set(sem_correspondencia)),
        "inventarioNormalizado": [
            {
                "chaveNormalizada": chave,
                "nomes": sorted({item.get("nomePadronizado") for item in itens if item.get("nomePadronizado")}),
                "preparos": sorted({item.get("preparo") for item in itens if item.get("preparo")}),
                "origens": sorted({item.get("origemDado") for item in itens if item.get("origemDado")}),
                "planos": sorted({plano for item in itens for plano in item.get("planosEmQueAparece", [])}),
                "ocorrencias": len(itens),
            }
            for chave, itens in sorted(por_nome.items())
        ],
    }
    return base, relatorio


def escrever(base: dict, relatorio: dict) -> None:
    SAIDA.write_text(json.dumps(base, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    inventario = {
        "geradoAntesDaBaseV1": True,
        "fontesDoInventario": [str(CATALOGO.relative_to(ROOT)), str(MODELO.relative_to(ROOT))],
        "ocorrenciasExtraidas": relatorio["ocorrenciasInventariadas"],
        "nomesNormalizadosUnicos": relatorio["nomesNormalizadosInventariados"],
        "gruposComDuplicatasOuVariantes": relatorio["duplicatasNormalizadas"],
        "alimentos": relatorio["inventarioNormalizado"],
    }
    INVENTARIO_JSON.write_text(json.dumps(inventario, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    INVENTARIO_MD.write_text(
        f"""# Inventário de alimentos antes da base nutricional v1\n\nGerado originalmente antes de alterar `public/data/alimentos.json` e reproduzido pelo mesmo pipeline da base.\n\n- Ocorrências extraídas: **{inventario['ocorrenciasExtraidas']}**\n- Nomes normalizados únicos: **{inventario['nomesNormalizadosUnicos']}**\n- Grupos com duplicatas ou variantes: **{inventario['gruposComDuplicatasOuVariantes']}**\n\nA normalização agrupa apenas grafia, acentos, pontuação e espaços. Preparos são preservados no inventário e não são considerados equivalentes por aproximação. O artefato completo está em `docs/alimentos-inventario-pre-base.json`.\n""",
        encoding="utf-8",
    )
    relatorio_publico = {chave: valor for chave, valor in relatorio.items() if chave != "inventarioNormalizado"}
    RELATORIO_JSON.write_text(json.dumps(relatorio_publico, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    pendentes = "\n".join(f"- {nome}" for nome in relatorio["alimentosAindaSemCorrespondencia"])
    texto = f"""# Relatório da base nutricional v1\n\n## Resultado\n\n- Alimentos adicionados: **{relatorio['alimentosGerados']}**\n- Registros disponíveis na TACO: **{relatorio['registrosTacoDisponiveis']}**\n- Nomes inventariados: **{relatorio['nomesNormalizadosInventariados']}**\n- Itens com peso documentado que podem ser resolvidos: **{relatorio['quantidadeItensComPesoDocumentadoResolvíveis']}**\n- Medidas caseiras documentadas incluídas: **{relatorio['medidasCaseirasDisponiveis']}**\n- Nomes ainda sem correspondência confiável: **{relatorio['quantidadeAindaSemCorrespondencia']}**\n\n## Fontes\n\nA composição nutricional usa exclusivamente a TACO 4ª edição. A TBCA não foi usada nesta v1 devido às condições publicadas para reprodução; consulte `data/fontes/README.md`. As medidas caseiras vêm somente dos pesos explícitos preservados no catálogo da nutricionista.\n\n## Pendências sem correspondência confiável\n\n{pendentes}\n\nO relatório estruturado completo está em `docs/alimentos-v1-relatorio.json`.\n"""
    RELATORIO_MD.write_text(texto, encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--check", action="store_true", help="falha se os artefatos versionados não forem reproduzíveis")
    args = parser.parse_args()
    base, relatorio = gerar()
    if args.check:
        esperado = json.loads(SAIDA.read_text(encoding="utf-8"))
        if esperado != base:
            raise SystemExit("public/data/alimentos.json está desatualizado; execute o gerador.")
    else:
        escrever(base, relatorio)
    print(json.dumps({"alimentos": len(base["alimentos"]), "pendentes": relatorio["quantidadeAindaSemCorrespondencia"], "medidas": relatorio["medidasCaseirasDisponiveis"]}, ensure_ascii=False))


if __name__ == "__main__":
    main()
